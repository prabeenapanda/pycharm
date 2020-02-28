import openpyxl as xl

#source File sheet
source_file = "c:\\Users\\prabeena.panda\\sourcee.xlsx"
target_file = "c:\\Users\\prabeena.panda\\targett.xlsx"
wb1 = xl.load_workbook(source_file)
sheet = wb1['Sheet1']

# columns source file
columns = ["Server IP" , "Description", "Group Name", "CPU", "Memory", "Storage"]

server_info = []
group_info=[]
i = 0
for row in sheet.rows:
    if i == 0:
        pass
    else:
        server_info.append([row[columns.index('Server IP')].value,
                            row[columns.index('CPU')].value,
                            row[columns.index('Memory')].value,
                            row[columns.index('Storage')].value,
                            row[columns.index('Description')].value])

        group_info.append([row[columns.index('Server IP')].value,
                           row[columns.index('Group Name')].value])
    i +=1
print(server_info)

Owb= xl.load_workbook(target_file)
Ows = Owb.worksheets[0]

for row_num, row in enumerate(server_info):
    for col_num , value in enumerate(row):
        Ows.cell(row=row_num+3, column=col_num+1).value = value

for row_num, row in enumerate(group_info):
    for col_num , value in enumerate(row):
        Ows.cell(row=row_num+3, column=col_num+8).value = value

Owb.save(target_file)





