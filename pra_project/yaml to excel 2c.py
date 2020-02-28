import yaml
import re
import openpyxl as xl

with open("C:\\Users\\prabeena.panda\\Downloads\\cgs.yaml") as file:
    documents = yaml.full_load(file)
col = ["Group Name" , "Description", "Storage Location", "Storage","Priority"," Datastore",
       "|Vms","Replication","Server IP","Virtual Server Name"]


vms=documents["Virtual Server Name"]
groups = documents["Group Name"]

target_file = "c:\\Users\\prabeena.panda\\targe.xlsx"
Owb= xl.load_workbook(target_file)
Ows = Owb.worksheets[0]



for row_num, group in enumerate(groups):
    for vm in vms[row_num].split(","):
        vm_details = vm.split("__##__")
        Ows.cell(row=row_num+3, column=9).value = vm_details[1]
        Ows.cell(row=row_num + 3, column=10).value = vm_details[0]



Owb.save(target_file)
