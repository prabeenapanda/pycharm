import openpyxl as xl

filename = "c:\\Users\\prabeena.panda\\sourcee.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

filename1 = "c:\\Users\\prabeena.panda\\targett.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

sr = ws1.max_row
sc = ws1.max_column
tr=ws2.max_row
tc=ws2.max_column


for j in range (1,tc+1):
     c = ws2.cell(row=2, column=j)
     for l in range (1,sc+1):
          d= ws1.cell(row=1,column=l)
          if c.value == d.value:
                 for p in range(2, sr + 1):
                        for r in range(1, sc + 1):
                              #e = ws1.cell(row=p, column=r)
                              for m in range(3, tr + 1):
                                     for n in range(1, tc + 1):
                                           ws2.cell(row=m,column=n).value = ws1.cell(row=p, column=r).value

wb2.save(str(filename1))

